"""Validation and duplicate removal for discovered companies."""

import csv
import re
from pathlib import Path
from typing import Iterator, Optional
from urllib.parse import urlparse

from agents.base_agent import BaseClass
from config.geography import canonical_city, matches_requested_city
from config.targeting import MIN_EMPLOYEE_COUNT, MIN_TURNOVER_CR, match_target_industry, parse_turnover_range
from schemas.company import Company


BACKEND_DIR = Path(__file__).resolve().parent.parent
DEFAULT_EXISTING_DATA_DIR = BACKEND_DIR / "Existing-data"
MASTER_EXPORT_PATH = BACKEND_DIR / "exports" / "All_Extracted_Leads.xlsx"


class ValidationAgent(BaseClass):
    """Removes duplicates and rejects only records with known ICP failures.

    Public directories rarely expose financials or headcount, so missing data is
    marked as unverified instead of silently excluding an otherwise useful lead.
    """

    def execute(
        self,
        companies: list[Company],
        *,
        existing_excel_path: Optional[str] = None,
        requested_location: Optional[str] = None,
    ) -> list[Company]:
        source = Path(existing_excel_path) if existing_excel_path else DEFAULT_EXISTING_DATA_DIR
        existing_keys = self._load_existing_keys(source)
        if MASTER_EXPORT_PATH.exists():
            existing_keys.update(self._load_existing_keys(MASTER_EXPORT_PATH))
        target_city = canonical_city(requested_location) or (requested_location or "").strip()
        kept: list[Company] = []
        seen: set[str] = set()

        for company in companies:
            keys = self._company_keys(company)
            if not keys:
                continue
            if keys & seen:
                continue
            seen.update(keys)

            if keys & existing_keys:
                continue

            if target_city:
                if not matches_requested_city(company.city, company.address, requested_location):
                    continue

            notes = self._validation_notes(company)
            rejected = any(note.startswith("rejected:") for note in notes)
            status = "rejected" if rejected else ("validated" if not notes else "unverified")
            if not rejected:
                kept.append(company.model_copy(update={
                    "validation_status": status,
                    "validation_notes": notes,
                    "remarks": company.remarks or "; ".join(notes),
                }))

        return kept

    def _validation_notes(self, company: Company) -> list[str]:
        notes: list[str] = []
        if company.industry:
            if not match_target_industry(company.industry):
                notes.append("rejected: declared industry is outside target taxonomy")
        else:
            notes.append("unverified: industry not supplied")

        if company.turnover:
            min_cr, max_cr = parse_turnover_range(company.turnover)
            if min_cr is None:
                notes.append("unverified: turnover could not be parsed")
            elif max_cr is not None and max_cr < MIN_TURNOVER_CR:
                notes.append(f"rejected: turnover below {MIN_TURNOVER_CR} Cr")
            elif min_cr < MIN_TURNOVER_CR:
                # Slab straddles the cutoff (e.g. "5 Cr to 25 Cr" vs a 10 Cr
                # minimum) - GST lookups only ever return a range, never an
                # exact figure, so this can't be resolved automatically.
                notes.append(f"unverified: turnover slab '{company.turnover}' straddles {MIN_TURNOVER_CR} Cr threshold - needs manual review")
        else:
            notes.append("unverified: turnover unavailable")

        employees = self._number(company.employee_count)
        if company.employee_count:
            if employees is None:
                notes.append("unverified: employee count could not be parsed")
            elif employees < MIN_EMPLOYEE_COUNT:
                notes.append(f"rejected: employee count below {MIN_EMPLOYEE_COUNT}")
        else:
            notes.append("unverified: employee count unavailable")
        return notes

    @staticmethod
    def _number(value: Optional[str]) -> Optional[float]:
        if not value:
            return None
        match = re.search(r"\d+(?:[,.]\d+)?", value.replace(",", ""))
        return float(match.group()) if match else None

    @staticmethod
    def _company_keys(company: Company) -> set[str]:
        keys: set[str] = set()
        if company.gst:
            normalized_gst = re.sub(r"\W+", "", company.gst).upper()
            keys.add(f"gst:{normalized_gst}")
        if company.website:
            parsed = urlparse(company.website if "://" in company.website else f"https://{company.website}")
            host = parsed.netloc.lower().removeprefix("www.")
            if host and "tradeindia.com" not in host:
                keys.add(f"website:{host}")
        if company.email:
            keys.add(f"email:{company.email.strip().lower()}")
        for number in (company.phone, company.phone_alt):
            if not number:
                continue
            digits = re.sub(r"\D+", "", number)
            if len(digits) >= 7:
                keys.add(f"phone:{digits[-10:]}")
        name = re.sub(r"\W+", "", company.company_name).lower()
        if name:
            keys.add(f"name:{name}")
        return keys

    def _load_existing_keys(self, path: Path) -> set[str]:
        if not path.exists():
            return set()
        files = [path] if path.is_file() else [*path.rglob("*.csv"), *path.rglob("*.xlsx")]
        keys: set[str] = set()
        for file_path in files:
            for company in self._read_existing_companies(file_path):
                keys.update(self._company_keys(company))
        return keys

    def _read_existing_companies(self, path: Path) -> Iterator[Company]:
        if path.suffix.lower() == ".csv":
            for encoding in ("utf-8-sig", "cp1252", "latin-1"):
                try:
                    with path.open("r", encoding=encoding, newline="") as handle:
                        rows = list(csv.DictReader(handle))
                    break
                except UnicodeDecodeError:
                    continue
            else:
                return
            for row in rows:
                yield self._company_from_row(row)
            return

        if path.suffix.lower() != ".xlsx":
            return
        try:
            from openpyxl import load_workbook
        except ImportError as exc:
            raise RuntimeError("Excel validation requires openpyxl; install dependencies first.") from exc

        workbook = load_workbook(path, read_only=True, data_only=True)
        try:
            sheet = workbook.active
            rows = sheet.iter_rows(values_only=True)
            headers = next(rows, ())
            for row in rows:
                yield self._company_from_row({str(headers[i] or ""): row[i] for i in range(min(len(headers), len(row)))})
        finally:
            workbook.close()

    def _company_from_row(self, row: dict) -> Company:
        values = {re.sub(r"[^a-z0-9]", "", str(key).lower()): str(value).strip() if value is not None else "" for key, value in row.items()}
        def get(*names: str) -> Optional[str]:
            return next((values[name] for name in names if values.get(name)), None)
        return Company(
            company_name=get("companyname", "company", "name") or "",
            gst=get("gst", "gstin"), city=get("city"),
            website=get("websiteurl", "website", "domain"),
            email=get("emailid", "email"),
            phone=get("mobilenumber", "contactnumber", "phone", "mobile"),
            phone_alt=get("alternatemobilenumber", "alternatephone", "altphone"),
        )

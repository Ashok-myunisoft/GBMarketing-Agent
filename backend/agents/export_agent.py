
from datetime import datetime
from pathlib import Path
from typing import Optional

from agents.base_agent import BaseClass
from schemas.company import Company

BACKEND_DIR = Path(__file__).resolve().parent.parent
MASTER_EXPORT_PATH = BACKEND_DIR / "exports" / "All_Extracted_Leads.xlsx"

EXPORT_COLUMNS = [
    ("Company Name", "company_name"), ("GST", "gst"), ("Turn Over", "turnover"),
    ("Region", "region"), ("City", "city"), ("Industry Type", "industry"),
    ("Contact Person", "contact_person"), ("Designation", "designation"),
    ("Mobile Number", "phone"), ("Alternate Mobile Number", "phone_alt"),
    ("Email ID", "email"), ("LinkedIN Id", "linkedin_url"),
    ("Website URL", "website"), ("Remarks", "remarks"), ("Followup", "followup"),
]
LEGACY_GST_HEADERS = {"GST Confidence", "GST Sources"}

class ExportAgent(BaseClass):
    def execute(self, companies: list[Company], output_path: Optional[str] = None) -> str:
        try:
            from openpyxl import Workbook
            from openpyxl.styles import Font
        except ImportError as exc:
            raise RuntimeError("Excel export requires openpyxl; install dependencies first.") from exc

        path = Path(output_path) if output_path else MASTER_EXPORT_PATH
        path.parent.mkdir(parents=True, exist_ok=True)
        headers = [header for header, _ in EXPORT_COLUMNS]
        if path.exists():
            from openpyxl import load_workbook
            workbook = load_workbook(path)
            sheet = workbook["Leads"] if "Leads" in workbook.sheetnames else workbook.active
            existing_headers = [cell.value for cell in sheet[1]]
            if existing_headers != headers:
                legacy_columns = [i for i, header in enumerate(existing_headers, start=1) if header in LEGACY_GST_HEADERS]
                remaining_headers = [header for header in existing_headers if header not in LEGACY_GST_HEADERS]
                if legacy_columns and remaining_headers == headers:
                    for column_index in sorted(legacy_columns, reverse=True):
                        sheet.delete_cols(column_index)
                else:
                    raise ValueError(f"Master export has unexpected headers: {path}")
        else:
            workbook = Workbook()
            sheet = workbook.active
            sheet.title = "Leads"
            sheet.append(headers)
            for cell in sheet[1]:
                cell.font = Font(bold=True)
        for company in companies:
            sheet.append([
                ("; ".join(value) if isinstance(value, list) else value)
                if (value := getattr(company, field)) is not None else ""
                for _, field in EXPORT_COLUMNS
            ])
        sheet.freeze_panes = "A2"
        sheet.auto_filter.ref = sheet.dimensions
        for column in sheet.columns:
            letter = column[0].column_letter
            sheet.column_dimensions[letter].width = min(max(len(str(cell.value or "")) for cell in column) + 2, 50)
        workbook.save(path)
        return str(path.resolve())
